#python3

import openpyxl
import copy

wb = openpyxl.load_workbook("Template.xlsx")
sheet = wb.active

new_wb = openpyxl.Workbook()

new_sheet = new_wb.create_sheet("1")
num = 1

for row in sheet.values :
    for i, content in enumerate(row) :
        cell = sheet.cell(num, i+1)
        new_cell = new_sheet.cell(num, i+1)
        new_cell.value = content
        new_cell.font = copy.copy(cell.font)
        new_cell.fill = copy.copy(cell.fill)
        if cell.comment :
            new_cell.commet = cell.comment
    num = num + 1

new_wb.save("to_4.xlsx")